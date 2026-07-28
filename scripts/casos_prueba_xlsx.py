#!/usr/bin/env python3
"""
Genera la PLANILLA DE EJECUCIÓN de QA (`CASOS_PRUEBA_QA.xlsx`) a partir de
`CASOS_PRUEBA_QA.md`.

El .md es la fuente de verdad: al agregar o cambiar un caso allí, se vuelve a correr este
script y la planilla queda al día. Así no hay dos listas que mantener en paralelo (que es
como terminan divergiendo y QA prueba una versión vieja).

Uso:  python3 scripts/casos_prueba_xlsx.py
Salida: CASOS_PRUEBA_QA.xlsx en la raíz del repo.
"""
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = Path(__file__).resolve().parents[1]
ORIGEN = RAIZ / 'CASOS_PRUEBA_QA.md'
DESTINO = RAIZ / 'CASOS_PRUEBA_QA.xlsx'

FUENTE = 'Arial'
AZUL = '16233F'
CYAN = '00C3FF'

# Prioridad: el .md la marca con emoji; en la planilla se filtra mejor como texto.
PRIORIDAD = {'🔴': 'Alta', '🟡': 'Media', '🟢': 'Baja'}

# Encabezado de sección: "## 11. Constructor de correos HTML  ✅"
RE_SECCION = re.compile(r'^##\s+(?:\d+\.\s*)?(.+?)\s*$')
# Fila de caso: "| CP-HTML-01 | 🔴 | Caso | Pasos | Resultado esperado |".
# El sufijo opcional de letra cubre los CP intercalados a mano (CP-PAY-07b).
RE_CASO = re.compile(r'^\|\s*(CP-[A-Z0-9]+-\d+[a-z]?)\s*\|(.+)\|\s*$')

ESTADOS = ['Pasó', 'No pasó', 'Bloqueado', 'No aplica']


def limpiar(texto: str) -> str:
    """Quita el marcado del .md que no aporta en una celda de Excel."""
    t = texto.strip()
    t = re.sub(r'\*\*(.+?)\*\*', r'\1', t)          # negritas
    t = re.sub(r'`(.+?)`', r'\1', t)                # código
    t = re.sub(r'\s{2,}', ' ', t)                    # espacios dobles que deja quitar emoji
    return t.strip()


def leer_casos(md: str):
    """Devuelve [(seccion, id, prioridad, caso, pasos, esperado)] en orden del documento."""
    casos = []
    seccion = ''
    for linea in md.splitlines():
        m_sec = RE_SECCION.match(linea)
        if m_sec:
            titulo = m_sec.group(1)
            # Fuera los emoji de estado y las notas entre paréntesis del título.
            titulo = re.sub(r'[✅🧩]', '', titulo)
            seccion = limpiar(titulo).strip()
            continue
        m_caso = RE_CASO.match(linea)
        if not m_caso:
            continue
        cp_id = m_caso.group(1)
        # `-1` porque la fila termina en "|": el split deja un campo vacío al final.
        campos = [limpiar(c) for c in m_caso.group(2).split('|')[:-1]]
        while len(campos) < 4:
            campos.append('')
        prioridad = PRIORIDAD.get(campos[0].strip(), campos[0].strip())
        casos.append((seccion, cp_id, prioridad, campos[1], campos[2], campos[3]))
    return casos


COLUMNAS = [
    ('ID', 12),
    ('Módulo', 34),
    ('Prioridad', 11),
    ('Caso', 40),
    ('Pasos', 52),
    ('Resultado esperado', 62),
    ('Resultado obtenido', 34),
    ('Estado', 13),
    ('Observaciones / defecto', 30),
    ('Probado por', 16),
    ('Fecha', 13),
]


def construir(casos):
    wb = Workbook()

    # ── Hoja de ejecución ──
    ws = wb.active
    ws.title = 'Casos'

    borde = Border(*[Side(style='thin', color='D6DEE8')] * 4)
    cab_fill = PatternFill('solid', fgColor=AZUL)
    cab_font = Font(name=FUENTE, bold=True, color='FFFFFF', size=10)

    for i, (titulo, ancho) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=1, column=i, value=titulo)
        c.fill = cab_fill
        c.font = cab_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = borde
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[1].height = 28

    zebra = PatternFill('solid', fgColor='F4F7FB')
    for fila, (seccion, cp_id, prioridad, caso, pasos, esperado) in enumerate(casos, start=2):
        valores = [cp_id, seccion, prioridad, caso, pasos, esperado, '', '', '', '', '']
        for col, valor in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=valor)
            c.font = Font(name=FUENTE, size=10)
            c.border = borde
            c.alignment = Alignment(
                vertical='top',
                wrap_text=col in (2, 4, 5, 6, 7, 9),
                horizontal='center' if col in (1, 3, 8, 11) else 'left',
            )
            if fila % 2 == 0:
                c.fill = zebra
        # Las columnas que llena QA se marcan en amarillo: se ve de un vistazo qué se edita.
        for col in (7, 8, 9, 10, 11):
            ws.cell(row=fila, column=col).fill = PatternFill('solid', fgColor='FFFDE7')

    ultima = len(casos) + 1
    ws.auto_filter.ref = f'A1:K{ultima}'
    ws.freeze_panes = 'B2'

    # Desplegable de estado: evita que cada quien escriba "OK"/"ok"/"PASO" y no se pueda contar.
    dv = DataValidation(type='list', formula1='"{}"'.format(','.join(ESTADOS)), allow_blank=True)
    dv.error = 'Elige un valor de la lista.'
    dv.promptTitle = 'Estado del caso'
    dv.prompt = 'Pasó / No pasó / Bloqueado / No aplica'
    ws.add_data_validation(dv)
    dv.add(f'H2:H{ultima}')

    dvf = DataValidation(type='date', operator='greaterThan', formula1='DATE(2020,1,1)', allow_blank=True)
    ws.add_data_validation(dvf)
    dvf.add(f'K2:K{ultima}')

    # ── Hoja de resumen ──
    rs = wb.create_sheet('Resumen')
    rs['A1'] = 'Resumen de ejecución — Casos de prueba QA MailConnect'
    rs['A1'].font = Font(name=FUENTE, bold=True, size=14, color=AZUL)
    rs['A2'] = 'Se calcula solo desde la hoja "Casos". No escribir aquí.'
    rs['A2'].font = Font(name=FUENTE, size=9, italic=True, color='6B7A90')

    rs['A4'] = 'Total de casos'
    rs['B4'] = f'=COUNTA(Casos!A2:A{ultima})'
    encabezados_estado = [(5 + i, e) for i, e in enumerate(ESTADOS)]
    for fila, estado in encabezados_estado:
        rs.cell(row=fila, column=1, value=estado)
        rs.cell(row=fila, column=2, value=f'=COUNTIF(Casos!$H$2:$H${ultima},A{fila})')
        rs.cell(row=fila, column=3, value=f'=IFERROR(B{fila}/$B$4,0)').number_format = '0.0%'
    rs['A9'] = 'Sin ejecutar'
    rs['B9'] = f'=$B$4-SUM(B5:B8)'
    rs['C9'] = '=IFERROR(B9/$B$4,0)'
    rs['C9'].number_format = '0.0%'

    rs['A11'] = 'Por prioridad'
    rs['A11'].font = Font(name=FUENTE, bold=True, size=11, color=AZUL)
    for i, p in enumerate(['Alta', 'Media', 'Baja']):
        f = 12 + i
        rs.cell(row=f, column=1, value=p)
        rs.cell(row=f, column=2, value=f'=COUNTIF(Casos!$C$2:$C${ultima},A{f})')
        rs.cell(row=f, column=3, value=f'=COUNTIFS(Casos!$C$2:$C${ultima},A{f},Casos!$H$2:$H${ultima},"No pasó")')
    rs['C11'] = 'Con fallo'
    rs['C11'].font = Font(name=FUENTE, bold=True, size=9, color=AZUL)

    rs['A16'] = 'Por módulo'
    rs['A16'].font = Font(name=FUENTE, bold=True, size=11, color=AZUL)
    for col, titulo in enumerate(['Módulo', 'Casos', 'Pasó', 'No pasó', 'Sin ejecutar'], start=1):
        c = rs.cell(row=17, column=col, value=titulo)
        c.font = Font(name=FUENTE, bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor=AZUL)

    modulos = []
    for seccion, *_ in casos:
        if seccion not in modulos:
            modulos.append(seccion)
    for i, mod in enumerate(modulos):
        f = 18 + i
        rs.cell(row=f, column=1, value=mod)
        rs.cell(row=f, column=2, value=f'=COUNTIF(Casos!$B$2:$B${ultima},$A{f})')
        rs.cell(row=f, column=3, value=f'=COUNTIFS(Casos!$B$2:$B${ultima},$A{f},Casos!$H$2:$H${ultima},"Pasó")')
        rs.cell(row=f, column=4, value=f'=COUNTIFS(Casos!$B$2:$B${ultima},$A{f},Casos!$H$2:$H${ultima},"No pasó")')
        rs.cell(row=f, column=5, value=f'=B{f}-COUNTIFS(Casos!$B$2:$B${ultima},$A{f},Casos!$H$2:$H${ultima},"<>")')

    for fila in rs.iter_rows(min_row=1, max_row=17 + len(modulos), max_col=5):
        for c in fila:
            if c.font.size is None or c.font.name != FUENTE:
                c.font = Font(name=FUENTE, size=10, bold=c.font.bold, color=c.font.color)
    rs.column_dimensions['A'].width = 40
    for col in 'BCDE':
        rs.column_dimensions[col].width = 14

    # ── Hoja de instrucciones ──
    ins = wb.create_sheet('Cómo usarla')
    texto = [
        ('Cómo usar esta planilla', True),
        ('', False),
        ('1. Ejecuta los casos de la hoja "Casos" en orden o filtrando por módulo/prioridad.', False),
        ('2. Llena SOLO las columnas amarillas: Resultado obtenido, Estado, Observaciones,', False),
        ('   Probado por y Fecha. Las demás vienen del documento y no se editan aquí.', False),
        ('3. "Estado" es un desplegable: Pasó / No pasó / Bloqueado / No aplica.', False),
        ('   Usa "Bloqueado" cuando el caso no se pudo ejecutar por una dependencia', False),
        ('   (por ejemplo, un endpoint que todavía no está desplegado — ver DESPLIEGUE.md).', False),
        ('4. La hoja "Resumen" se calcula sola a partir de la columna Estado. Si al abrir', False),
        ('   el archivo el resumen sale en blanco, pulsa F9 (recalcular): las fórmulas se', False),
        ('   escriben sin valores cacheados y Excel las evalúa al abrir.', False),
        ('', False),
        ('Ojo: esta planilla se GENERA desde CASOS_PRUEBA_QA.md.', True),
        ('Si se agregan casos al documento y se vuelve a generar, lo que hayas escrito aquí', False),
        ('se pierde. Guarda una copia con la fecha antes de regenerar, o pásale los', False),
        ('resultados a quien la regenere.', False),
        ('', False),
        ('Prioridades: Alta = crítico (bloquea salida a producción) · Media · Baja.', False),
    ]
    for i, (linea, negrita) in enumerate(texto, start=1):
        c = ins.cell(row=i, column=1, value=linea)
        c.font = Font(name=FUENTE, size=11, bold=negrita, color=AZUL if negrita else '000000')
    ins.column_dimensions['A'].width = 100
    ins.sheet_properties.tabColor = CYAN

    return wb


def main():
    if not ORIGEN.exists():
        print(f'No se encontró {ORIGEN}', file=sys.stderr)
        return 1
    casos = leer_casos(ORIGEN.read_text(encoding='utf-8'))
    if not casos:
        print('No se encontró ningún caso de prueba en el documento.', file=sys.stderr)
        return 1
    construir(casos).save(DESTINO)
    modulos = len({c[0] for c in casos})
    print(f'{DESTINO.name}: {len(casos)} casos de prueba en {modulos} módulos.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
