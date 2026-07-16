'''
Lambda para generación de reportes de campañas de email.
Diseñada para ser invocada como herramienta (tool) de un agente AWS Bedrock.

Acciones soportadas:
  - status_summary      : Cantidades agrupadas por estado para una campaña
  - full_report         : Todos los registros enviados con su estado actual
  - open_rate           : Porcentajes de apertura/clic/rebote de una o varias campañas
  - campaign_comparison : Comparación de métricas entre dos campañas
  - list_campaigns      : Listado de campañas del cliente

Parámetros de entrada requeridos (via event['parameters']):
  action            str  (requerido) Acción a ejecutar
  customerId        str  (requerido) ID del cliente
  customerName      str  (requerido) Nombre interno del cliente en DynamoDB
  campaignId        str  (requerido para: status_summary, full_report, campaign_comparison)
  campaignId2       str  (requerido para: campaign_comparison)
  campaignIds       str  (requerido para: open_rate) IDs separados por coma
  page              int  (opcional) Página para full_report (default 1)
  pageSize          int  (opcional) Registros por página en full_report (default 500, max 1000)
'''
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from decimal import Decimal
from collections import defaultdict
from typing import List, Dict, Optional
from datetime import datetime, timedelta, timezone

import boto3
from boto3.dynamodb.conditions import Key, Attr


REGION = 'us-east-1'
dynamodb = boto3.resource('dynamodb', region_name=REGION)

table_campaign = dynamodb.Table("campaign")
table_process = dynamodb.Table("process")

# Mapeo de códigos numéricos a nombres legibles
STATE_NAMES = {
    1: 'Enviado',
    2: 'Entregado',
    3: 'Rechazado',
    4: 'Abierto',
    5: 'Clicado',
    6: 'Rebote',
    7: 'Queja',
    8: 'FallaRenderizado',
    9: 'Retrasado',
    10: 'Suscrito',
}

# Prioridad para determinar el estado "actual" de un email.
STATE_PRIORITY = {
    1: 1,   # Enviado
    9: 2,   # Retrasado
    8: 3,   # FallaRenderizado
    3: 4,   # Rechazado
    2: 5,   # Entregado
    6: 6,   # Rebote
    10: 7,  # Suscrito
    7: 8,   # Queja
    4: 9,   # Abierto
    5: 10,  # Clicado
}


# ─────────────────────────────────────────────────
# Utilidades de DynamoDB
# ─────────────────────────────────────────────────

def decimal_to_python(obj):
    """Convierte tipos Decimal de DynamoDB a int/float nativos de Python."""
    if isinstance(obj, Decimal):
        return int(obj) if obj % 1 == 0 else float(obj)
    raise TypeError(f"Object of type {type(obj)} is not JSON serializable")


def query_gsi_with_pagination(table, gsi_name: str, 
                                key_condition: dict, 
                                projection: Optional[str] = None) -> List[Dict]:
    """
    Realiza Query en un GSI con manejo automático de paginación.
    
    Args:
        table_name: Nombre de la tabla DynamoDB
        gsi_name: Nombre del GSI
        key_condition: {'pk': value, 'sk': value} o {'pk': value}
        projection: Expression de proyección (ej: 'id, name, #dt')
    
    Returns:
        Lista de items que cumplen la condición
    """
    print(f"[QUERY_GSI] GSI: {gsi_name}, Keys: {list(key_condition.keys())}")
    
    try:
        
        # Construir KeyConditionExpression
        key_attrs = list(key_condition.keys())
        if len(key_attrs) == 1:
            pk = key_attrs[0]
            key_expr = Key(pk).eq(key_condition[pk])
        else:
            pk, sk = key_attrs[0], key_attrs[1]
            key_expr = Key(pk).eq(key_condition[pk]) & Key(sk).eq(key_condition[sk])
        
        kwargs = {
            'IndexName': gsi_name,
            'KeyConditionExpression': key_expr
        }
        if projection:
            kwargs['ProjectionExpression'] = projection
            # Agregar mapping para palabras reservadas
            kwargs['ExpressionAttributeNames'] = {
                '#dt': 'date'
            }

        items = []
        response = table.query(**kwargs)
        items.extend(response.get('Items', []))
        count = response.get('Count', 0)
        scanned = response.get('ScannedCount', 0)
        print(f"[QUERY_GSI] Primera página: {count} items, ScannedCount: {scanned}")
        
        page_count = 1
        while 'LastEvaluatedKey' in response:
            page_count += 1
            response = table.query(ExclusiveStartKey=response['LastEvaluatedKey'], **kwargs)
            page_items = response.get('Items', [])
            items.extend(page_items)
            count = response.get('Count', 0)
            print(f"[QUERY_GSI] Página {page_count}: {count} items")
        
        print(f"[QUERY_GSI] ✓ Total: {len(items)} items en {page_count} páginas")
        return items
        
    except Exception as exc:
        print(f"[QUERY_GSI] ✗ ERROR: {str(exc)}")
        raise


def query_table_by_key(table: str, key_name: str, key_value: str,
                        projection: Optional[str] = None) -> Optional[Dict]:
    """
    Realiza Query por clave primaria (directa, sin GSI).
    
    Args:
        table_name: Nombre de la tabla
        key_name: Nombre de la clave primaria
        key_value: Valor de la clave
        projection: Expression de proyección
    
    Returns:
        Item encontrado o None
    """
    print(f"[QUERY_KEY] Clave: {key_name}={key_value}")
    
    try:
        
        kwargs = {
            'KeyConditionExpression': Key(key_name).eq(key_value)
        }
        if projection:
            kwargs['ProjectionExpression'] = projection

        response = table.query(**kwargs)
        items = response.get('Items', [])
        
        if items:
            print(f"[QUERY_KEY] ✓ Item encontrado")
            return items
        
        print(f"[QUERY_KEY] ✗ Item no encontrado")
        return None
        
    except Exception as exc:
        print(f"[QUERY_KEY] ✗ ERROR: {str(exc)}")
        raise


def query_status_by_process(customer_name: str, process_id: str) -> List[Dict]:
    """Estados de UN proceso desde la tabla única {customer}_sendStatus (PK processId).
    Reemplaza el scan de la antigua tabla-por-proceso {customer}_sendStatus_{proceso}."""
    table = dynamodb.Table(f'{customer_name}_sendStatus')
    items: List[Dict] = []
    kwargs = {'KeyConditionExpression': Key('processId').eq(process_id)}
    try:
        while True:
            resp = table.query(**kwargs)
            items.extend(resp.get('Items', []))
            last_key = resp.get('LastEvaluatedKey')
            if not last_key:
                break
            kwargs['ExclusiveStartKey'] = last_key
    except Exception as exc:
        print(f"[QUERY_STATUS] ⚠ {customer_name}_sendStatus proceso {process_id}: {exc}")
    return items


def query_detail_by_process(customer_name: str, process_id: str) -> List[Dict]:
    """Detalle de UN proceso desde la tabla única {customer}_sendDetail (PK processId +
    SK sendDetailId). Reemplaza el scan de la antigua tabla-por-proceso
    {customer}_sendDetail_{proceso}. Pagina por LastEvaluatedKey."""
    table = dynamodb.Table(f'{customer_name}_sendDetail')
    items: List[Dict] = []
    kwargs = {'KeyConditionExpression': Key('processId').eq(process_id)}
    try:
        while True:
            resp = table.query(**kwargs)
            items.extend(resp.get('Items', []))
            last_key = resp.get('LastEvaluatedKey')
            if not last_key:
                break
            kwargs['ExclusiveStartKey'] = last_key
    except Exception as exc:
        print(f"[QUERY_DETAIL] ⚠ {customer_name}_sendDetail proceso {process_id}: {exc}")
    return items


def scan_full_table_with_pagination(table_name: str,
                                     projection: Optional[str] = None) -> List[Dict]:
    """
    Escanea una tabla completa (SOLO para tablas pequeñas como sendStatus).
    
    Args:
        table_name: Nombre de la tabla
        projection: Expression de proyección
    
    Returns:
        Lista de todos los items
    """
    print(f"[SCAN_FULL] Tabla: {table_name}")
    
    try:
        table = dynamodb.Table(table_name)
        
        kwargs = {}
        if projection:
            kwargs['ProjectionExpression'] = projection

        items = []
        response = table.scan(**kwargs)
        items.extend(response.get('Items', []))
        print(f"[SCAN_FULL] Primera página: {len(items)} items")
        
        page_count = 1
        while 'LastEvaluatedKey' in response:
            page_count += 1
            response = table.scan(ExclusiveStartKey=response['LastEvaluatedKey'], **kwargs)
            page_items = response.get('Items', [])
            items.extend(page_items)
        
        print(f"[SCAN_FULL] ✓ Total: {len(items)} items en {page_count} páginas")
        return items
        
    except Exception as exc:
        print(f"[SCAN_FULL] ✗ ERROR: {str(exc)}")
        raise


def get_current_state_per_message(send_status_records: List[Dict]) -> Dict:
    """
    Procesa historial de estados y devuelve el estado más relevante por messageId.
    
    Args:
        send_status_records: Lista de registros de estado
    
    Returns:
        Dict con {messageId: {'state': int, 'date': str}}
    """
    print(f"[GET_CURRENT_STATE] Procesando {len(send_status_records)} registros de estado")
    
    states: Dict = {}
    invalid_records = 0
    
    for idx, record in enumerate(send_status_records):
        try:
            msg_id = record.get('messageId', '')
            if not msg_id:
                invalid_records += 1
                continue
            
            state_num = int(record.get('state', 0))
            date = record.get('date', '')

            if msg_id not in states:
                states[msg_id] = {'state': state_num, 'date': date}
            else:
                current_priority = STATE_PRIORITY.get(states[msg_id]['state'], 0)
                new_priority = STATE_PRIORITY.get(state_num, 0)
                if new_priority > current_priority:
                    states[msg_id] = {'state': state_num, 'date': date}
        
        except Exception as exc:
            print(f"[GET_CURRENT_STATE] ⚠ Error en registro {idx}: {str(exc)}")
            invalid_records += 1
    
    print(f"[GET_CURRENT_STATE] ✓ {len(states)} mensajes únicos, {invalid_records} inválidos ignorados")
    return states


# ─────────────────────────────────────────────────
# Funciones de negocio - Con Query + GSI
# ─────────────────────────────────────────────────

def get_campaign(campaign_id: str, customer_id: str, customer_name: str) -> Dict:
    """
    Obtiene una campaña específica del cliente usando GSI.
    
    GSI esperado: campaign -> customerId-campaignId-index
    
    Args:
        campaign_id: ID de la campaña
        customer_id: ID del cliente (para validación)
        customer_name: Nombre del cliente
    
    Returns:
        Dict con los datos de la campaña
    
    Raises:
        ValueError: Si campaña no existe o no pertenece al cliente
    """
    print(f"[GET_CAMPAIGN] Buscando campaña: {campaign_id} para cliente: {customer_name}")
    
    try:
        items = query_table_by_key(
            table_campaign,
            'campaignId',
            campaign_id,
            "campaignName,campaignState,channel,customerId"
        )
        
        print(f"[GET_CAMPAIGN] Items encontrados: {len(items)}")
        
        if not items:
            raise ValueError(f"Campaña no encontrada: {campaign_id}")
        
        campaign = items[0]
        
        # Validación adicional
        if campaign.get('customerId') != customer_id:
            print(f"[GET_CAMPAIGN] ✗ Campaña no pertenece al cliente especificado")
            raise ValueError(f"Campaña {campaign_id} no pertenece al cliente {customer_name}")
        
        print(f"[GET_CAMPAIGN] ✓ Campaña encontrada: {campaign.get('campaignName', 'N/A')}")
        return campaign
        
    except Exception as exc:
        print(f"[GET_CAMPAIGN] ✗ ERROR: {str(exc)}")
        raise


def get_processes_for_campaign(campaign_id: str) -> List[Dict]:
    """
    Obtiene todos los procesos de una campaña usando GSI.
    
    GSI esperado: process -> campaignId-processId-index
    
    Args:
        campaign_id: ID de la campaña
    
    Returns:
        Lista de procesos
    """
    print(f"[GET_PROCESSES] Buscando procesos para campaña: {campaign_id}")
    
    try:
        processes = query_gsi_with_pagination(
            table_process,
            'campaignId-date',
            {'campaignId': campaign_id},
            projection='processId, campaignId, campaignName, #dt, registersToSend, registersOnSpool, quantityBlacklist, quantityUnsubscribe, quantityDeletions'
        )        
        print(f"[GET_PROCESSES] ✓ {len(processes)} procesos encontrados")
        return processes
        
    except Exception as exc:
        print(f"[GET_PROCESSES] ✗ ERROR: {str(exc)}")
        raise


def collect_campaign_metrics(campaign_id: str, customer_name: str, 
                              customer_id: str) -> Dict:
    """
    Recopila métricas agregadas de todos los procesos de una campaña.
    
    Args:
        campaign_id: ID de la campaña
        customer_name: Nombre interno del cliente
        customer_id: ID del cliente (para validación)
    
    Returns:
        Dict con campaign, totals y stateCounts
    """
    print(f"[COLLECT_METRICS] Iniciando para campaña: {campaign_id}, cliente: {customer_name}")
    
    try:
        # Validar que la campaña existe y pertenece al cliente
        campaign = get_campaign(campaign_id, customer_id, customer_name)
        
        processes = get_processes_for_campaign(campaign_id)
        print(f"[COLLECT_METRICS] {len(processes)} procesos encontrados")
        
        if not processes:
            print(f"[COLLECT_METRICS] ⚠ No hay procesos para esta campaña")
            return {
                'campaign': campaign,
                'totals': {
                    'registersOnSpool': 0,
                    'registersToSend': 0,
                    'excludedBlacklist': 0,
                    'excludedUnsubscribe': 0,
                    'excludedInvalidEmail': 0,
                },
                'stateCounts': {},
            }

        total_on_spool = 0
        total_to_send = 0
        total_blacklist = 0
        total_unsubscribe = 0
        total_deletions = 0
        state_counts: Dict = defaultdict(int)

        for process in processes:
            # Omitir procesos de muestras
            if "-Samples" in  process['campaignName']:
                print("Saltando...")
                continue
            process_id = process['processId']
            
            try:
                on_spool = int(process.get('registersOnSpool', 0))
                to_send = int(process.get('registersToSend', 0))
                blacklist = int(process.get('quantityBlacklist', 0))
                unsub = int(process.get('quantityUnsubscribe', 0))
                delet = int(process.get('quantityDeletions', 0))
                
                total_on_spool += on_spool
                total_to_send += to_send
                total_blacklist += blacklist
                total_unsubscribe += unsub
                total_deletions += delet
                
                print(f"[COLLECT_METRICS] {process_id}: toSend={to_send}, onSpool={on_spool}")

                # Estados de este proceso desde la tabla única {customer}_sendStatus.
                print(f"[COLLECT_METRICS] Consultando estados del proceso {process_id}")
                try:
                    status_records = query_status_by_process(customer_name, process_id)
                    print(f"[COLLECT_METRICS] {len(status_records)} registros de estado")
                    
                    current_states = get_current_state_per_message(status_records)
                    
                    for msg_state in current_states.values():
                        state_num = msg_state['state']
                        state_counts[state_num] += 1
                    
                except Exception as exc:
                    print(f"[COLLECT_METRICS] ⚠ No se pudo procesar estados para {process_id}: {str(exc)}")
                    print(f"[COLLECT_METRICS] La tabla podría no existir si aún no se han procesado envíos")
                    
            except Exception as exc:
                print(f"[COLLECT_METRICS] ✗ Error en proceso {process_id}: {str(exc)}")
                continue

        print(f"[COLLECT_METRICS] ✓ Métricas recopiladas: {sum(state_counts.values())} estados")
        
        return {
            'campaign': campaign,
            'totals': {
                'registersOnSpool': total_on_spool,
                'registersToSend': total_to_send,
                'excludedBlacklist': total_blacklist,
                'excludedUnsubscribe': total_unsubscribe,
                'excludedInvalidEmail': total_deletions,
            },
            'stateCounts': dict(state_counts),
        }
        
    except Exception as exc:
        print(f"[COLLECT_METRICS] ✗ ERROR CRÍTICO: {str(exc)}")
        raise


# ─────────────────────────────────────────────────
# REPORTES
# ─────────────────────────────────────────────────

def generar_excel(records):
    #items
    '''
    {
            'campaignId': campaign_id,
            'campaignName': campaign_name,
            'totalRecords': total_records,
            'page': page,
            'pageSize': page_size,
            'totalPages': total_pages,
            'records': page_records,
        }
    '''
    items = records['records']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Envíos"
    
    # Headers
    headers = ['Nombre', 'Email', 'Estado', 'Fecha Envío']
    ws.append(headers)
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Datos (neutraliza inyección de fórmulas: prefija ' si empieza con = + - @)
    def _safe(value):
        s = '' if value is None else str(value)
        return ("'" + s) if (s and s[0] in ('=', '+', '-', '@', '\t', '\r')) else s

    for item in items:
        ws.append([
            _safe(item.get('nombre', '')),
            _safe(item['email']),
            _safe(item['currentState']),
            _safe(item['sendDate'])
        ])
    
    # Auto-ajustar columnas
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Guardar Excel en BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)  # ⚠️ IMPORTANTE: volver al inicio

    return excel_buffer.getvalue()

def report_status_summary(campaign_id: str, customer_name: str, 
                          customer_id: str) -> Dict:
    """
    Resumen de cantidades por estado para una campaña.
    
    Retorna:
        Dict con estado de la campaña y conteos por estado
    """
    print(f"[REPORT_STATUS_SUMMARY] Generando para: {campaign_id}")
    
    try:
        metrics = collect_campaign_metrics(campaign_id, customer_name, customer_id)
        campaign = metrics['campaign']
        
        states_summary = {
            STATE_NAMES.get(k, f'Estado_{k}'): v
            for k, v in metrics['stateCounts'].items()
        }
        
        result = {
            'campaignId': campaign_id,
            'campaignName': campaign.get('campaignName', ''),
            'campaignDate': campaign.get('date', ''),
            'campaignState': campaign.get('campaignState', ''),
            'exclusions': metrics['totals'],
            'statesSummary': states_summary,
            'grandTotalTracked': sum(metrics['stateCounts'].values()),
        }
        
        print(f"[REPORT_STATUS_SUMMARY] ✓ Total rastreado: {result['grandTotalTracked']}")
        return result
        
    except Exception as exc:
        print(f"[REPORT_STATUS_SUMMARY] ✗ ERROR: {str(exc)}")
        raise


def report_full_records(campaign_id: str, customer_name: str, customer_id: str,
                         page: int = 1, page_size: int = 500) -> Dict:
    """
    Reporte completo de registros enviados con su estado actual.
    Soporta paginación para campañas grandes.
    
    Args:
        campaign_id: ID de la campaña
        customer_name: Nombre interno del cliente
        customer_id: ID del cliente
        page: Número de página (default 1)
        page_size: Registros por página (default 500, máximo 1000)
    
    Returns:
        Dict con registros paginados
    """
    print(f"[REPORT_FULL_RECORDS] Generando para: {campaign_id}, página: {page}, pageSize: {page_size}")
    
    try:
        # Validar campaña
        campaign = get_campaign(campaign_id, customer_id, customer_name)
        campaign_name = campaign.get('campaignName', '')
        processes = get_processes_for_campaign(campaign_id)
        print(f"[REPORT_FULL_RECORDS] {len(processes)} procesos encontrados")
        
        if not processes:
            print(f"[REPORT_FULL_RECORDS] ✗ No hay procesos para esta campaña")
            return {
                'error': 'No se encontraron procesos para esta campaña',
                'campaignId': campaign_id,
                'records': []
            }

        all_records = []
        processed_details = 0
        failed_processes = 0

        for proc_idx, process in enumerate(processes):
            if "-Samples" in  process['campaignName']:
                print("Saltando...")
                continue
            process_id = process['processId']
            print(f"[REPORT_FULL_RECORDS] Procesando {proc_idx + 1}/{len(processes)}: {process_id}")

            try:
                # Tabla ÚNICA de detalle: Query por processId (antes: scan de la tabla por proceso).
                print(f"[REPORT_FULL_RECORDS] Consultando detalle del proceso {process_id}")
                send_details = query_detail_by_process(customer_name, process_id)
                print(f"[REPORT_FULL_RECORDS] {len(send_details)} detalles obtenidos")

                # Estados de este proceso desde la tabla única {customer}_sendStatus.
                try:
                    print(f"[REPORT_FULL_RECORDS] Consultando estados del proceso {process_id}")
                    status_records = query_status_by_process(customer_name, process_id)
                    print(f"[REPORT_FULL_RECORDS] {len(status_records)} estados obtenidos")
                    
                    current_states = get_current_state_per_message(status_records)
                except Exception as exc:
                    print(f"[REPORT_FULL_RECORDS] ⚠ No se pudo obtener estados: {str(exc)}")
                    current_states = {}

                # Construir registros
                for detail in send_details:
                    try:
                        send_detail_id = detail.get('sendDetailId', '')
                        state_info = current_states.get(send_detail_id, {})
                        state_num = state_info.get('state', 1)

                        record = {
                            'uniqueId': detail.get('uniqueId', ''),
                            'email': detail.get('email', ''),
                            'sendDate': detail.get('date', ''),
                            'currentState': STATE_NAMES.get(state_num, f'Estado_{state_num}'),
                            'stateCode': state_num,
                            'lastStateDate': state_info.get('date', ''),
                            'processId': process_id,
                        }
                        all_records.append(record)
                        processed_details += 1
                        
                    except Exception as exc:
                        print(f"[REPORT_FULL_RECORDS] ⚠ Error en detalle: {str(exc)}")
                        continue

            except Exception as exc:
                print(f"[REPORT_FULL_RECORDS] ✗ Error en proceso: {str(exc)}")
                failed_processes += 1
                continue

        print(f"[REPORT_FULL_RECORDS] ✓ {processed_details} registros procesados, {failed_processes} procesos fallidos")

        # Paginación
        page_size = min(max(page_size, 1), 1000)
        total_records = len(all_records)
        start = (page - 1) * page_size
        end = start + page_size
        page_records = all_records[start:end]
        
        total_pages = max(1, (total_records + page_size - 1) // page_size)
        print(f"[REPORT_FULL_RECORDS] Página {page} de {total_pages}, {len(page_records)} registros")

        return (campaign_name,{
            'campaignId': campaign_id,
            'campaignName': campaign_name,
            'totalRecords': total_records,
            'page': page,
            'pageSize': page_size,
            'totalPages': total_pages,
            'records': page_records,
        })
        
    except Exception as exc:
        print(f"[REPORT_FULL_RECORDS] ✗ ERROR CRÍTICO: {str(exc)}")
        raise


def report_open_rate(campaign_ids: List[str], customer_name: str, 
                      customer_id: str) -> Dict:
    """
    Calcula tasas de apertura, clic, entrega, rebote y queja
    para una o varias campañas del cliente.
    
    Args:
        campaign_ids: Lista de IDs de campañas
        customer_name: Nombre interno del cliente
        customer_id: ID del cliente
    
    Returns:
        Dict con métricas de cada campaña
    """
    print(f"[REPORT_OPEN_RATE] Generando para {len(campaign_ids)} campaña(s)")
    
    results = []

    for camp_idx, campaign_id in enumerate(campaign_ids):
        print(f"[REPORT_OPEN_RATE] {camp_idx + 1}/{len(campaign_ids)}: {campaign_id}")
        
        try:
            metrics = collect_campaign_metrics(campaign_id, customer_name, customer_id)
            campaign = metrics['campaign']
            state_counts = metrics['stateCounts']
            total_sent = metrics['totals']['registersToSend']

            print(f"[REPORT_OPEN_RATE] {campaign_id}: totalSent={total_sent}")

            total_delivered = state_counts.get(2, 0)
            total_opened = state_counts.get(4, 0)
            total_clicked = state_counts.get(5, 0)
            total_bounced = state_counts.get(6, 0)
            total_complained = state_counts.get(7, 0)
            total_rejected = state_counts.get(3, 0)
            total_delayed = state_counts.get(9, 0)

            # Para entrega efectiva: abierto o clicado también es entregado
            effective_delivered = total_delivered + total_opened + total_clicked

            def pct(n):
                return f'{round(n / total_sent * 100, 2)}%' if total_sent > 0 else '0%'

            campaign_result = {
                'campaignId': campaign_id,
                'campaignName': campaign.get('campaignName', ''),
                'campaignDate': campaign.get('date', ''),
                'metrics': {
                    'totalSent': total_sent,
                    'totalDelivered': effective_delivered,
                    'totalOpened': total_opened,
                    'totalClicked': total_clicked,
                    'totalBounced': total_bounced,
                    'totalComplained': total_complained,
                    'totalRejected': total_rejected,
                    'totalDelayed': total_delayed,
                    'deliveryRate': pct(effective_delivered),
                    'openRate': pct(total_opened),
                    'clickRate': pct(total_clicked),
                    'bounceRate': pct(total_bounced),
                    'complaintRate': pct(total_complained),
                },
            }
            results.append(campaign_result)
            print(f"[REPORT_OPEN_RATE] ✓ openRate={campaign_result['metrics']['openRate']}")
            
        except Exception as exc:
            print(f"[REPORT_OPEN_RATE] ✗ Error en {campaign_id}: {str(exc)}")
            continue

    print(f"[REPORT_OPEN_RATE] ✓ {len(results)}/{len(campaign_ids)} campañas procesadas")
    return {'campaigns': results}


def report_campaign_comparison(campaign_id_1: str, campaign_id_2: str,
                                customer_name: str, customer_id: str) -> Dict:
    """
    Compara métricas de dos campañas del cliente.
    campaign_id_1 = base (anterior), campaign_id_2 = actual.
    
    Args:
        campaign_id_1: ID de la campaña base
        campaign_id_2: ID de la campaña actual
        customer_name: Nombre interno del cliente
        customer_id: ID del cliente
    
    Returns:
        Dict con comparación de métricas
    """
    print(f"[REPORT_COMPARISON] Comparando {campaign_id_1} vs {campaign_id_2}")
    
    try:
        data = report_open_rate([campaign_id_1, campaign_id_2], customer_name, customer_id)
        campaigns = data.get('campaigns', [])

        if len(campaigns) < 2:
            print(f"[REPORT_COMPARISON] ✗ No se pudieron obtener ambas campañas")
            raise ValueError('No se pudieron obtener datos de ambas campañas')

        c1, c2 = campaigns[0], campaigns[1]
        m1, m2 = c1['metrics'], c2['metrics']

        print(f"[REPORT_COMPARISON] Base: {c1['campaignName']}")
        print(f"[REPORT_COMPARISON] Actual: {c2['campaignName']}")

        def rate_diff(key):
            r1 = float(m1[key].replace('%', ''))
            r2 = float(m2[key].replace('%', ''))
            diff = round(r2 - r1, 2)
            sign = '+' if diff >= 0 else ''
            return {
                'campaign1': m1[key],
                'campaign2': m2[key],
                'difference': f'{sign}{diff}%',
                'trend': 'sube' if diff > 0 else ('baja' if diff < 0 else 'igual'),
            }

        def abs_diff(key):
            d = m2[key] - m1[key]
            return {
                'campaign1': m1[key],
                'campaign2': m2[key],
                'difference': d,
                'trend': 'sube' if d > 0 else ('baja' if d < 0 else 'igual'),
            }

        open_rate_diff = rate_diff('openRate')

        result = {
            'campaign1': c1,
            'campaign2': c2,
            'comparison': {
                'sent': abs_diff('totalSent'),
                'deliveryRate': rate_diff('deliveryRate'),
                'openRate': open_rate_diff,
                'clickRate': rate_diff('clickRate'),
                'bounceRate': rate_diff('bounceRate'),
                'complaintRate': rate_diff('complaintRate'),
            },
            'summary': (
                f"Campaña '{c2['campaignName']}' vs '{c1['campaignName']}': "
                f"apertura {open_rate_diff['campaign2']} "
                f"({open_rate_diff['difference']} respecto a la anterior). "
                f"Entrega: {rate_diff('deliveryRate')['campaign2']}, "
                f"Rebote: {rate_diff('bounceRate')['campaign2']}."
            ),
        }
        
        print(f"[REPORT_COMPARISON] ✓ Comparación generada")
        return result
        
    except Exception as exc:
        print(f"[REPORT_COMPARISON] ✗ ERROR: {str(exc)}")
        raise


def report_list_campaigns(customer_id: str, customer_name: str, 
                          limit: int = 5, campaign_type: str = '',
                          include_record_count: bool = False) -> Dict:
    """
    Lista las campañas del cliente con información clave usando GSI.
    
    GSI esperado: campaign -> customerId-campaignId-index
    
    Args:
        customer_id: ID del cliente
        customer_name: Nombre interno del cliente
        limit: Máximo de campañas a retornar (default 5, max 100)
        campaign_type: Filtrar por tipo (ej: "promocion", "newsletter")
        include_record_count: Incluir total de registros (más lento)
    
    Returns:
        Dict con lista de campañas del cliente
    """
    print(f"[LIST_CAMPAIGNS] Buscando campañas del cliente: {customer_name}")
    print(f"[LIST_CAMPAIGNS] Limit: {limit}, Type filter: {campaign_type}")
    
    try:
        # Validar límite
        limit = min(max(int(limit), 1), 100)
        print(f"[LIST_CAMPAIGNS] Límite validado: {limit}")
        
        # Obtener todas las campañas del cliente usando GSI
        print(f"[LIST_CAMPAIGNS] Consultando GSI customerId-campaignId-index...")
        all_campaigns = query_gsi_with_pagination(
            table_campaign,
            'customerId-date',
            {'customerId': customer_id},
            projection='campaignId, campaignName, campaignState, channel, #dt, description'
        )
        print(f"[LIST_CAMPAIGNS] Campañas encontradas: {len(all_campaigns)}")
        
        # Filtrar por tipo si se especifica
        if campaign_type:
            campaign_type_lower = campaign_type.lower()
            filtered = []
            for campaign in all_campaigns:
                camp_type = campaign.get('campaignType', '').lower()
                camp_name = campaign.get('campaignName', '').lower()
                if campaign_type_lower in camp_type or campaign_type_lower in camp_name:
                    filtered.append(campaign)
            print(f"[LIST_CAMPAIGNS] Después de filtrar por tipo: {len(filtered)}")
            all_campaigns = filtered
        
        total_campaigns = len(all_campaigns)
        campaigns_to_return = all_campaigns[:limit]
        
        # Construir respuesta
        result_campaigns = []
        for idx, campaign in enumerate(campaigns_to_return):
            print(f"[LIST_CAMPAIGNS] Procesando {idx + 1}/{len(campaigns_to_return)}")
            
            try:
                campaign_id = campaign.get('campaignId', '')
                
                campaign_info = {
                    'campaignName': campaign.get('campaignName', 'Sin nombre'),
                    'campaignState': campaign.get('campaignState', ''),
                    'campaignType': campaign.get('campaignType', 'Sin especificar'),
                    'campaignDate': campaign.get('date', ''),
                    'description': campaign.get('description', ''),
                    '_campaignId': campaign_id,
                }
                
                # Opcional: agregar total de registros
                if include_record_count:
                    try:
                        processes = get_processes_for_campaign(campaign_id)
                        total_records = sum(int(p.get('registersToSend', 0)) for p in processes)
                        campaign_info['totalRecords'] = total_records
                        print(f"[LIST_CAMPAIGNS] {campaign_id}: {total_records} registros")
                    except Exception as exc:
                        print(f"[LIST_CAMPAIGNS] ⚠ No se pudo obtener registros para {campaign_id}: {str(exc)}")
                        campaign_info['totalRecords'] = 0
                
                result_campaigns.append(campaign_info)
                
            except Exception as exc:
                print(f"[LIST_CAMPAIGNS] ⚠ Error procesando campaña: {str(exc)}")
                continue
        
        print(f"[LIST_CAMPAIGNS] ✓ {len(result_campaigns)} campañas retornadas")
        
        return {
            'customerId': customer_id,
            'customerName': customer_name,
            'totalCampaigns': total_campaigns,
            'campaignsReturned': len(result_campaigns),
            'limitApplied': limit,
            'filterApplied': campaign_type if campaign_type else 'ninguno',
            'campaigns': result_campaigns,
        }
        
    except Exception as exc:
        print(f"[LIST_CAMPAIGNS] ✗ ERROR: {str(exc)}")
        raise

def generate_presigned_url(bucket_name, data, filename):
    print("Cargando a S3")
    s3_client = boto3.client('s3')
    key = f'reportes/{filename}'
    
    # Guardar datos en S3
    s3_client.put_object(
        Bucket=bucket_name,
        Key=key,
        Body=data,
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    print("Prefirmando url")
    # Generar URL prefirmada
    url = s3_client.generate_presigned_url(
        'get_object',
        Params={'Bucket': bucket_name, 'Key': key},
        ExpiresIn=3600
    )
    return url

# ─────────────────────────────────────────────────
# Handler principal
# ─────────────────────────────────────────────────

def lambda_handler(event: dict, context: dict) -> dict:
    """
    Punto de entrada de la Lambda.
    Invocación como function tool de AWS Bedrock Agent.

    Formato del event (Bedrock Agent):
    {
      "actionGroup": "...",
      "function":    "...",
      "parameters":  [
        {"name": "action",       "type": "string", "value": "status_summary"},
        {"name": "customerId",   "type": "string", "value": "..."},
        {"name": "customerName",  "type": "string", "value": "..."},
        {"name": "campaignId",   "type": "string", "value": "..."}
      ]
    }
    """
    print("═" * 80)
    print("[LAMBDA] ✸ INICIANDO")
    print("═" * 80)

    print("[DEBUG] Event completo:")
    print(event)
    
    response_body = '{}'

    try:
        # Extraer parámetros
        print("[LAMBDA] Extrayendo parámetros...")
        
        raw_params = event.get('parameters', [])
        params: dict = {}
        for p in raw_params:
            params[p['name']] = p['value']
        print(f"[LAMBDA] {params}")

        # Parámetros requeridos
        action = params.get('action', '').strip().lower()
        customer_id = params.get('customerId', '').strip()
        customer_name = params.get('customerName', '').strip()

        print(f"[LAMBDA] Action: {action}")
        print(f"[LAMBDA] CustomerId: {customer_id}")
        print(f"[LAMBDA] CustomerName: {customer_name}")

        bucket_name = f"{customer_name}.document"

        # Validar parámetros requeridos
        if not action or not customer_id or not customer_name:
            print("[LAMBDA] ✗ Parámetros requeridos faltantes")
            return {
                'response': {
                    'actionGroup': event.get('actionGroup', ''),
                    'function': event.get('function', ''),
                    'functionResponse': {
                        'responseBody': {
                            'TEXT': {
                                'body': json.dumps({
                                    'error': 'Parámetros requeridos: action, customerId, customerName'
                                }, ensure_ascii=False)
                            }
                        }
                    }
                }
            }

        result: dict = {}

        # ── get customer data ───────────────────────────────────────────────
        if action == 'get_customer_data':
            print("[LAMBDA] Ejecutando: get_customer_data")
            #result = get_customer_data(customer_id)
       
        # ── status_summary ──────────────────────────────────────────────────
        if action == 'status_summary':
            print("[LAMBDA] Ejecutando: status_summary")
            campaign_id = params.get('campaignId', '').strip()
            if not campaign_id:
                result = {'error': 'Parámetro requerido: campaignId'}
            else:
                result = report_status_summary(campaign_id, customer_name, customer_id)

        # ── full_report ─────────────────────────────────────────────────────
        elif action == 'full_report':
            print("[LAMBDA] Ejecutando: full_report")
            campaign_name = ""
            formatted_date = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
            campaign_id = params.get('campaignId', '').strip()
            if not campaign_id:
                result = {'error': 'Parámetro requerido: campaignId'}
            else:
                page = int(params.get('page', 1))
                page_size = int(params.get('pageSize', 500))
                campaign_name, data = report_full_records(campaign_id, customer_name, customer_id, page, page_size)
            excel_data = generar_excel(data)
            type_send = params.get('typeSend', '')
            print(f"Tipo de envio: {type_send}")
            if (type_send == "Url"):
                # Generar url prefirmada con el reporte
                report_name = f"ReporteEnvios_{campaign_name}_{formatted_date}.xlsx"
                result = generate_presigned_url(bucket_name, excel_data, report_name)                
            else:
                pass
        # ── open_rate ───────────────────────────────────────────────────────
        elif action == 'open_rate':
            print("[LAMBDA] Ejecutando: open_rate")
            raw_ids = params.get('campaignIds', params.get('campaignId', ''))
            campaign_ids = [c.strip() for c in raw_ids.split(',') if c.strip()]
            
            if not campaign_ids:
                result = {'error': 'Parámetro requerido: campaignIds (o campaignId)'}
            else:
                result = report_open_rate(campaign_ids, customer_name, customer_id)

        # ── campaign_comparison ─────────────────────────────────────────────
        elif action == 'campaign_comparison':
            print("[LAMBDA] Ejecutando: campaign_comparison")
            campaign_id_1 = params.get('campaignId', '').strip()
            campaign_id_2 = params.get('campaignId2', '').strip()
            
            if not campaign_id_1 or not campaign_id_2:
                result = {'error': 'Parámetros requeridos: campaignId y campaignId2'}
            else:
                result = report_campaign_comparison(campaign_id_1, campaign_id_2, customer_name, customer_id)

        # ── list_campaigns ──────────────────────────────────────────────────
        elif action == 'get_campaigns_list':
            print("[LAMBDA] Ejecutando: get_campaigns_list")
            limit = int(params.get('limit', 5))
            campaign_type = params.get('campaignType', '').strip()
            include_count = params.get('includeRecordCount', 'false').lower() == 'true'
            result = report_list_campaigns(customer_id, customer_name, limit, campaign_type, include_count)

        # ── acción desconocida ──────────────────────────────────────────────
        else:
            print(f"[LAMBDA] ✗ Acción no reconocida: {action}")
            result = {
                'error': f"Acción no reconocida: '{action}'",
                'validActions': [
                    'status_summary',
                    'full_report',
                    'open_rate',
                    'campaign_comparison',
                    'get_campaigns_list',
                ],
            }

        print("[LAMBDA] Preparando respuesta...")
        response_body = json.dumps(result, default=decimal_to_python, ensure_ascii=False)
        print(f"[LAMBDA] Respuesta: {len(response_body)} bytes")

    except Exception as exc:
        print(f"[LAMBDA] ✗ ERROR CRÍTICO: {str(exc)}")
        response_body = json.dumps(
            {'error': f'Error interno: {str(exc)}'},
            ensure_ascii=False
        )

    print("[LAMBDA] ✓ COMPLETADA")
    print(event.get('apiPath', ''))
    return {                
        'response': {
            'actionGroup': event.get('actionGroup', ''),
            'apiPath': event.get('apiPath', ''),
            'httpStatusCode': 200,
            'httpMethod': event.get('httpMethod', ''),
            'responseBody': {
                'TEXT': {
                    'body': response_body
                }
            }
        }
    }